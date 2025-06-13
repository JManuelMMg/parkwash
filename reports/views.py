from django.shortcuts import render
from django.views.generic import ListView, DetailView, CreateView, View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.contrib import messages
from .models import Report
from datetime import datetime, timedelta
from parking.models import Reservation
from django.contrib.auth import get_user_model
from django.db.models import Sum, Count, Avg, F
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML

User = get_user_model()

class StaffRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_staff

class ReportListView(LoginRequiredMixin, StaffRequiredMixin, ListView):
    model = Report
    template_name = 'reports/report_list.html'
    context_object_name = 'reports'

class ReportDetailView(LoginRequiredMixin, StaffRequiredMixin, DetailView):
    model = Report
    template_name = 'reports/report_detail.html'
    context_object_name = 'report'

class ReportCreateView(LoginRequiredMixin, StaffRequiredMixin, CreateView):
    model = Report
    template_name = 'reports/report_create.html'
    fields = ['title', 'report_type', 'start_date', 'end_date', 'content']
    success_url = reverse_lazy('reports:list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Reporte creado exitosamente.')
        return super().form_valid(form)

class DailyReportView(LoginRequiredMixin, StaffRequiredMixin, View):
    def get(self, request):
        today = datetime.now().date()
        context = self.generate_report(today, today)
        return render(request, 'reports/daily_report.html', context)

    def generate_report(self, start_date, end_date):
        # Implementar lógica de generación de reporte diario
        return {'start_date': start_date, 'end_date': end_date}

class WeeklyReportView(LoginRequiredMixin, StaffRequiredMixin, View):
    def get(self, request):
        today = datetime.now().date()
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
        context = self.generate_report(start_date, end_date)
        return render(request, 'reports/weekly_report.html', context)

    def generate_report(self, start_date, end_date):
        # Implementar lógica de generación de reporte semanal
        return {'start_date': start_date, 'end_date': end_date}

class MonthlyReportView(LoginRequiredMixin, StaffRequiredMixin, View):
    def get(self, request):
        today = datetime.now().date()
        start_date = today.replace(day=1)
        next_month = start_date.replace(day=28) + timedelta(days=4)
        end_date = next_month - timedelta(days=next_month.day)
        context = self.generate_report(start_date, end_date)
        return render(request, 'reports/monthly_report.html', context)

    def generate_report(self, start_date, end_date):
        # Implementar lógica de generación de reporte mensual
        return {'start_date': start_date, 'end_date': end_date}

class CustomReportView(LoginRequiredMixin, StaffRequiredMixin, View):
    def get(self, request):
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date and end_date:
            context = self.generate_report(start_date, end_date)
        else:
            context = {}
        return render(request, 'reports/custom_report.html', context)

    def generate_report(self, start_date, end_date):
        # Implementar lógica de generación de reporte personalizado
        return {'start_date': start_date, 'end_date': end_date}

class ExportReportView(LoginRequiredMixin, StaffRequiredMixin, View):
    def get(self, request, pk):
        report = Report.objects.get(pk=pk)
        # Implementar lógica de exportación de reporte
        return render(request, 'reports/export_report.html', {'report': report})

class ParkingStatsReportView(LoginRequiredMixin, StaffRequiredMixin, View):
    def get(self, request):
        # Filtros por fecha
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        reservations = Reservation.objects.filter(status__in=['COMPLETED', 'ACTIVE'])
        if start_date:
            reservations = reservations.filter(start_time__date__gte=start_date)
        if end_date:
            reservations = reservations.filter(end_time__date__lte=end_date)

        # Reporte general
        total_income = reservations.aggregate(total=Sum('total_cost'))['total'] or 0
        total_reservations = reservations.count()
        avg_duration = reservations.annotate(
            duration=F('end_time') - F('start_time')
        ).aggregate(avg=Avg('duration'))['avg']
        most_used_spaces = reservations.values('space__space_number').annotate(
            count=Count('id')).order_by('-count')[:5]

        # Reporte por usuario
        user_stats = reservations.values('user__username').annotate(
            total_income=Sum('total_cost'),
            reservation_count=Count('id'),
        ).order_by('-total_income')

        context = {
            'total_income': total_income,
            'total_reservations': total_reservations,
            'avg_duration': avg_duration,
            'most_used_spaces': most_used_spaces,
            'user_stats': user_stats,
            'start_date': start_date,
            'end_date': end_date,
        }
        return render(request, 'reports/parking_stats_report.html', context)

class ExportParkingStatsPDFView(LoginRequiredMixin, StaffRequiredMixin, View):
    def get(self, request):
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        reservations = Reservation.objects.filter(status__in=['COMPLETED', 'ACTIVE'])
        if start_date:
            reservations = reservations.filter(start_time__date__gte=start_date)
        if end_date:
            reservations = reservations.filter(end_time__date__lte=end_date)

        total_income = reservations.aggregate(total=Sum('total_cost'))['total'] or 0
        total_reservations = reservations.count()
        avg_duration = reservations.annotate(
            duration=F('end_time') - F('start_time')
        ).aggregate(avg=Avg('duration'))['avg']
        most_used_spaces = reservations.values('space__space_number').annotate(
            count=Count('id')).order_by('-count')[:5]
        user_stats = reservations.values('user__username').annotate(
            total_income=Sum('total_cost'),
            reservation_count=Count('id'),
        ).order_by('-total_income')

        # Si tienes un logo, pon la ruta aquí (ejemplo: '/static/img/logo.png')
        logo_url = None
        context = {
            'total_income': total_income,
            'total_reservations': total_reservations,
            'avg_duration': avg_duration,
            'most_used_spaces': most_used_spaces,
            'user_stats': user_stats,
            'start_date': start_date,
            'end_date': end_date,
            'logo_url': logo_url,
        }
        html_string = render_to_string('reports/parking_stats_report_pdf.html', context)
        pdf_file = HTML(string=html_string).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=parking_report.pdf'
        return response

class ExportParkingStatsExcelView(LoginRequiredMixin, StaffRequiredMixin, View):
    def get(self, request):
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        reservations = Reservation.objects.filter(status__in=['COMPLETED', 'ACTIVE'])
        if start_date:
            reservations = reservations.filter(start_time__date__gte=start_date)
        if end_date:
            reservations = reservations.filter(end_time__date__lte=end_date)

        # User stats
        user_stats = reservations.values('user__username').annotate(
            total_income=Sum('total_cost'),
            reservation_count=Count('id'),
        ).order_by('-total_income')

        # Most used spaces
        most_used_spaces = reservations.values('space__space_number').annotate(
            count=Count('id')).order_by('-count')[:5]

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'Reporte General'
        ws1.append(['Espacio', 'Veces usado'])
        for space in most_used_spaces:
            ws1.append([space['space__space_number'], space['count']])

        ws2 = wb.create_sheet('Por Usuario')
        ws2.append(['Usuario', 'Ingresos', 'Reservas'])
        for user in user_stats:
            ws2.append([user['user__username'], float(user['total_income'] or 0), user['reservation_count']])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=parking_report.xlsx'
        wb.save(response)
        return response
